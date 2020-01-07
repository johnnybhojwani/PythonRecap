
from urllib.request import urlopen
from bs4 import BeautifulSoup
import openpyxl as xl
from openpyxl.styles import Font





webpage = 'https://www.boxofficemojo.com/weekend/chart/'


page = urlopen(webpage)			

soup = BeautifulSoup(page, 'html.parser')

title = soup.title

print(title.text)
##
##
##
##

#find all elements with the tag <table>
all_tables = soup.findAll('table')
##print(all_tables[0])

#since our table is the first table in the list of tables
movie_table = all_tables[0]

#within the movie table, find all rows
movie_rows = movie_table.findAll('tr')


#since we are interested in only the top 5 movies
#we go row by row
for x in range(1,6):
    # here we look for each cell (create a list of cells)
    td = movie_rows[x].findAll('td')
    print(td)
    print(td[0].text)   #number spot
    print(td[2].text)   #name of movie
    print(td[3].text)   #name of studio
    print(td[4].text)   #weekend gross 



    
##Write to an excel file


#create a new workbook
wb=  xl.Workbook()

#assign the active worksheet to 'ws' variable
ws = wb.active


#rename the worksheet
ws.title = 'Box Office Report'

#write headings to excel
ws['A1'] = 'No.'
ws['B1'] = 'Movie Title'
ws['C1'] = 'Studio'
ws['D1'] = 'Weekend Gross'
ws['E1'] = 'Total Gross'
ws['F1'] = 'Total weeks running'

##we just want the top 5 movies so we can
##reassign movie_rows to just those elements
##in the list

for x in range(1,6):
    td = movie_rows[x].findAll('td')
    print(td[0].text)
    tw = td[0].text
    title = td[2].text
    studio = td[3].text
    weekend_gross = td[4].text
    total_gross = td[9].text
    total_weeks = td[11].text


#write to specific cells using x as our row variable
    ws['A' + str(x+1)]= tw
    ws['B' + str(x+1)]= title
    ws['C' + str(x+1)]= studio
    ws['D' + str(x+1)]= weekend_gross
    ws['E' + str(x+1)]= total_gross
    ws['F' + str(x+1)]= total_weeks


#adjust column sizes
ws.column_dimensions['A'].width = 5
ws.column_dimensions['B'].width = 22
ws.column_dimensions['C'].width = 16
ws.column_dimensions['D'].width = 21
ws.column_dimensions['E'].width = 15
ws.column_dimensions['F'].width = 26


#apply formatting to the header row
header_font = Font(size=16,bold=True)

for cell in ws[1:1]:
    cell.font = header_font



#save the new workbook
wb.save('BoxOfficeReport.xlsx')
##    
##
##
##    
##    
##    
    





